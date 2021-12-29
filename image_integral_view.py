import random
import copy
import openpyxl


def print_matrix(matrix: list[list]):
    """
    Выводит матрицу
    :param matrix: матрица для вывода
    """
    for row in matrix:
        print(row)


def generate_random_matrix(width: int, height: int) -> list[list]:
    """
    :param width: ширина матрицы
    :param height: высота матрицы
    :return: матрица размерностью width*height,
    заполненная случайными цислами от 0 до 255
    """
    return [[random.randint(0, 255) for x in range(width)] for y in range(height)]


def integral_view(image: list[list]) -> list[list]:
    """
    :param image: матрица, заполненная числами типа int
    :return: интегральное представление исходной матрицы, с такой же размерностью
    """
    integral_matrix = copy.deepcopy(image)

    for y in range(len(integral_matrix)):
        for x in range(len(integral_matrix[0])):
            left_cell = integral_matrix[y][x - 1] if x > 0 else 0
            up_cell = integral_matrix[y - 1][x] if y > 0 else 0
            diagonal_cell = integral_matrix[y - 1][x - 1] if x > 0 and y > 0 else 0
            integral_matrix[y][x] += left_cell + up_cell - diagonal_cell

    return integral_matrix


def rect_sum(image: list[list], x1: int, y1: int, x2: int, y2: int) -> int:
    """
    Вычисляет сумму чисел внутри прямоугольника
    :param image: матрица, заполненная числами типа int
    :param x1: координата x левого верхнего угла прямоугольника
    :param y1: координата y левого верхнего угла прямоугольника
    :param x2: координата x правого нижнего угла прямоугольника
    :param y2: координата y правого нижнего угла прямоугольника
    :return: сумма чисел внутри прямоугольника
    """
    integral_matrix = integral_view(image)

    a_sum = integral_matrix[y1 - 1][x1 - 1] if y1 > 0 and x1 > 0 else 0
    b_sum = integral_matrix[y1 - 1][x2] if y1 > 0 else 0
    c_sum = integral_matrix[y2][x2]
    d_sum = integral_matrix[y2][x1 - 1] if x1 > 0 else 0

    return a_sum + c_sum - d_sum - b_sum


def get_xlsx_input() -> list[list]:
    """
    Считывает из файла типа .xslx матрицу
    :return: матрица
    """
    wb = openpyxl.load_workbook("xslx_input.xlsx")
    sheet = wb["Sheet1"]
    row_count = sheet.max_row + 1
    column_count = sheet.max_column + 1

    try:
        a = [[int(sheet.cell(x, y).value) for y in range(1, column_count)] for x in range(1, row_count)]
        return a
    except TypeError:
        quit("Что-то пошло не так, попробуйте заполнить целочисленными значениями пустые клетки")


def get_input():
    """
    Запрашивает у пользователя на ввод 4 координаты прямоугольника
    :return: координаты прямоугольника
    """
    u_input = None
    got_input = False
    input_message = 'Введите 4 целых числа >= 1 x1;y1;x2;y2 через ";" ' \
                    '\nгде x1, y1 - координаты левого верхнего угла прямоугольника, ' \
                    'а x2, y2 - координаты правого нижнего: '
    while not got_input:
        input_request = input(input_message)
        try:
            raw_input = [int(i) for i in input_request.strip().split(';') if i != '']
        except ValueError:
            print('Пример "1;2;3;4"')
            continue
        u_input = raw_input
        got_input = True
    return u_input


def main():
    """
    Основная логика
    """
    image = get_xlsx_input()
    coordinates = None
    while True:
        coordinates = get_input()

        passed_check = True
        for i in coordinates:
            if i < 1:
                passed_check = False
                print("Координаты должны быть > 1")
                break
        if not passed_check: continue

        if coordinates[0] > len(image[0]) or coordinates[2] > len(image[0]):
            print(f"x координыты должны быть <= {len(image[0])}")
            continue
        if coordinates[1] > len(image) or coordinates[3] > len(image):
            print(f"y координыты должны быть <= {len(image)}")
            continue
        break
    print("Изначальная матрица")
    print_matrix(image)
    print("Интегральная матрица")
    print_matrix(integral_view(image))

    user_rect_sum = rect_sum(image, coordinates[0] - 1, coordinates[1] - 1, coordinates[2] - 1, coordinates[3] -1)

    print(f'Сумма значений внутри прямоугольника c координатами '
          f'{coordinates[0]},{coordinates[1]},{coordinates[2]},{coordinates[3]} : {user_rect_sum}')


main()
